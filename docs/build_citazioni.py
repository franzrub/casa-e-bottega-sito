from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RUBY="7A1525"; CREAM="F3EDE4"; DARK="2B2B2B"; GREEN="1E7A3D"; AMBER="B8860B"
wb=Workbook()

# ---------- Sheet 1: NAP ----------
s=wb.active; s.title="NAP (da usare ovunque)"
s.sheet_view.showGridLines=False
title=Font(name="Arial",size=16,bold=True,color="FFFFFF")
s["A1"]="Casa e Bottega — dati NAP da usare IDENTICI su ogni portale"
s["A1"].font=title; s["A1"].fill=PatternFill("solid",start_color=RUBY)
s["A1"].alignment=Alignment(vertical="center")
s.merge_cells("A1:B1"); s.row_dimensions[1].height=30

rows=[
 ("Campo","Valore ESATTO (copia-incolla)"),
 ("Nome","Casa e Bottega"),
 ("Indirizzo","Via Gargano 13"),
 ("CAP / Città","71043 Manfredonia (FG)"),
 ("Regione / Paese","Puglia, Italia"),
 ("Telefono","⚠︎ da inserire — sempre lo STESSO numero ovunque"),
 ("Email","⚠︎ da inserire"),
 ("Sito web","https://www.casaebottegapuglia.it/"),
 ("Categoria","Casa vacanze / B&B (stile B&B)"),
 ("CIN (Cod. Id. Nazionale)","⚠︎ obbligatorio in Italia — inseriscilo dove richiesto (Booking, Airbnb, DMS)"),
]
hdr=Font(name="Arial",size=11,bold=True,color="FFFFFF")
cell=Font(name="Arial",size=11,color=DARK)
lab=Font(name="Arial",size=11,bold=True,color=DARK)
thin=Side(style="thin",color="D8CCBb"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
r=3
for i,(a,b) in enumerate(rows):
    s.cell(r,1,a); s.cell(r,2,b)
    if i==0:
        for c in (1,2):
            s.cell(r,c).font=hdr; s.cell(r,c).fill=PatternFill("solid",start_color=DARK)
    else:
        s.cell(r,1).font=lab; s.cell(r,1).fill=PatternFill("solid",start_color=CREAM)
        s.cell(r,2).font=cell
    for c in (1,2):
        s.cell(r,c).border=border; s.cell(r,c).alignment=Alignment(vertical="center",wrap_text=True)
    r+=1
s.column_dimensions["A"].width=24; s.column_dimensions["B"].width=70
note=("REGOLA D'ORO: le citazioni funzionano solo se Nome-Indirizzo-Telefono sono IDENTICI ovunque "
      "(stessa punteggiatura, stesse abbreviazioni). Meglio 12 citazioni coerenti che 100 sciatte. "
      "Non iscriverti a directory-spazzatura: possono danneggiare invece di aiutare.")
s.cell(r+1,1,note); s.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=2)
s.cell(r+1,1).font=Font(name="Arial",size=10,italic=True,color=RUBY)
s.cell(r+1,1).alignment=Alignment(wrap_text=True,vertical="top")
s.row_dimensions[r+1].height=60

# ---------- Sheet 2: Checklist ----------
c2=wb.create_sheet("Checklist citazioni")
c2.sheet_view.showGridLines=False
headers=["Portale","Categoria","Priorità","Costo","Link registrazione","Note","Stato"]
data=[
 # Priorità ALTA — gratis, massimo impatto
 ("Google Business Profile","Mappe / Ricerca","Alta","Gratis","business.google.com","La citazione n.1. Foto, orari, recensioni.","✅ Fatto"),
 ("Apple Business Connect","Mappe (Apple)","Alta","Gratis","businessconnect.apple.com","Apple Maps: tutti gli utenti iPhone. Spesso dimenticato.","Da fare"),
 ("Bing Places","Mappe / Ricerca","Alta","Gratis","bingplaces.com","Bing + Copilot. Puoi importare da Google.","Da fare"),
 ("TripAdvisor","Recensioni viaggi","Alta","Gratis","tripadvisor.it/Owners","Scheda struttura gratuita + recensioni.","Da fare"),
 ("DMS Puglia / ViaggiareinPuglia","Portale reg. ufficiale","Alta","Gratis / obbligo","dms.puglia.it","Obbligo regionale + citazione autorevole .puglia.it.","Verifica"),
 ("Booking.com","OTA","Alta","Commissione","admin.booking.com","Già attivo.","✅ Fatto"),
 ("Airbnb","OTA","Alta","Commissione","airbnb.it/host","Già attivo.","✅ Fatto"),
 # Priorità MEDIA — directory Italia + OTA
 ("PagineGialle","Directory IT","Media","Gratis (base)","paginegialle.it","Storica, buona autorità NAP in Italia.","Da fare"),
 ("PagineBianche","Directory IT","Media","Gratis","paginebianche.it","Stessa rete (Italiaonline).","Da fare"),
 ("Virgilio / Tuttocittà","Directory IT","Media","Gratis","tuttocitta.it","Scheda attività locale.","Da fare"),
 ("Yelp Italia","Recensioni locali","Media","Gratis","biz.yelp.it","Citazione + recensioni, alimenta Apple Maps.","Da fare"),
 ("Foursquare","Geodata","Media","Gratis","foursquare.com/venue/claim","Alimenta molte app di mappe di terze parti.","Da fare"),
 ("Expedia Group","OTA","Media","Commissione","join.expediagroup.com","Un accesso = Expedia + Hotels.com + Vrbo.","Valuta"),
 ("Vrbo / HomeToGo","OTA case vacanze","Media","Commissione","vrbo.com","Target case vacanze/famiglie.","Valuta"),
 ("Holidu","Aggregatore CV","Media","Commissione","holidu.it","Aggregatore case vacanze in crescita.","Valuta"),
 ("Bed-and-Breakfast.it","Directory B&B IT","Media","Freemium","bed-and-breakfast.it","Directory verticale storica in Italia.","Da fare"),
 ("BBPlanet","Directory B&B IT","Media","Freemium","bbplanet.it","13.500+ strutture, buona per il verticale.","Da fare"),
 # Priorità BASSA — territoriali / nicchia (backlink + local relevance)
 ("Gargano.it","Portale turistico locale","Bassa","Gratis/pagam.","gargano.it","Rilevanza locale + backlink territoriale.","Valuta"),
 ("IlGargano.com","Portale turistico locale","Bassa","Gratis/pagam.","ilgargano.com","Idem: contesto Gargano.","Valuta"),
 ("Cylex Italia","Directory business","Bassa","Gratis","cylex.it","Citazione NAP aggiuntiva.","Opzionale"),
 ("Hotfrog Italia","Directory business","Bassa","Gratis","hotfrog.it","Citazione NAP aggiuntiva.","Opzionale"),
 ("Misterimprese","Directory business IT","Bassa","Gratis","misterimprese.it","Citazione NAP aggiuntiva.","Opzionale"),
 ("Facebook (Pagina)","Social","Bassa","Gratis","facebook.com","Pagina con indirizzo = citazione + segnale social.","Da fare"),
 ("Instagram (business)","Social","Bassa","Gratis","instagram.com","Profilo business con tag luogo Manfredonia.","Da fare"),
]
# header
for j,h in enumerate(headers,1):
    cc=c2.cell(1,j,h); cc.font=hdr; cc.fill=PatternFill("solid",start_color=RUBY)
    cc.alignment=Alignment(horizontal="center",vertical="center"); cc.border=border
c2.row_dimensions[1].height=24
prio_fill={"Alta":"F4CCC C".replace(" ",""),"Media":"FCE5CD","Bassa":"EDEDED"}
for i,row in enumerate(data,2):
    for j,val in enumerate(row,1):
        cc=c2.cell(i,j,val); cc.font=cell; cc.border=border
        cc.alignment=Alignment(vertical="center",wrap_text=(j in(5,6)))
    # priority color
    pv=row[2]; c2.cell(i,3).fill=PatternFill("solid",start_color=prio_fill.get(pv,"FFFFFF"))
    c2.cell(i,3).alignment=Alignment(horizontal="center",vertical="center")
    # stato color
    st=row[6]
    if st.startswith("✅"):
        c2.cell(i,7).font=Font(name="Arial",size=11,bold=True,color=GREEN)
    elif st in("Da fare",):
        c2.cell(i,7).font=Font(name="Arial",size=11,bold=True,color=AMBER)
    c2.cell(i,7).alignment=Alignment(horizontal="center",vertical="center")
    # zebra
    if i%2==0:
        for j in (1,2,4,5,6):
            c2.cell(i,j).fill=PatternFill("solid",start_color="FBF7F1")
widths=[30,22,10,16,32,46,12]
for j,w in enumerate(widths,1):
    c2.column_dimensions[get_column_letter(j)].width=w
c2.freeze_panes="A2"
c2.auto_filter.ref=f"A1:G{len(data)+1}"

wb.save("Casa-e-Bottega-citazioni-directory.xlsx")
print("saved")
